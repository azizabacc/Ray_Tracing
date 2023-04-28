# Python program to read the map (excel) and generate the list of walls

import openpyxl
from wall import Wall
import matplotlib.pyplot as plt


def create_wall():
    # Give the location of the file to open the workbook 
    # path ="piece_rectangulaire.xlsx"
    path = "1mur.xlsx"

    # workbook object is created and activeted
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # number of rows of the excel sheet
    row_count = sheet_obj.max_row
    # Creation of the list of walls
    wall_list = []
    # filling the list of walls
    for value in sheet_obj.iter_rows(min_row=2, max_row=row_count, min_col=1, max_col=6, values_only=True):
        wall_list.append(Wall(value[4], [(value[0], value[1]), (value[2], value[3])], value[5]))
    # print(len(wall_list))
    # for i in range(0,len(wall_list)):
    # print(wall_list[i].point_list)

    return wall_list


def plot_wall(wall_list):
    # plots the walls
    for i in range(0, len(wall_list)):
        j = 0
        #
        while j < len(wall_list[i].point_list) - 1:
            print(len(wall_list[i].point_list))
            # print(wall_list[i].point_list[j][0], wall_list[i].point_list[j][1])
            #print([wall_list[i].point_list[j][0], wall_list[i].point_list[j + 1][0]], [wall_list[i].point_list[j][1], wall_list[i].point_list[j + 1][1]])
            plt.plot([wall_list[i].point_list[j][0], wall_list[i].point_list[j + 1][0]],
                     [wall_list[i].point_list[j][1], wall_list[i].point_list[j + 1][1]], c='black')
            j = j + 2


"""def plot_wall(wall_list):
    #plots the walls
    for i in range(0, len(wall_list)):
        for j in range(0,len(wall_list[i].point_list)-1) :
            #print(len(wall_list[i].point_list))
            #print(wall_list[i].point_list[j][0], wall_list[i].point_list[j][1])
            print([wall_list[i].point_list[j][0], wall_list[i].point_list[j+1][0]],
                     [wall_list[i].point_list[j][1], wall_list[i].point_list[j+1][1]])
            plt.plot([wall_list[i].point_list[j][0], wall_list[i].point_list[j+1][0]],
                     [wall_list[i].point_list[j][1], wall_list[i].point_list[j+1][1]], c='black')"""
