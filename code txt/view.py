# Python program to read the map (excel) and generate the list of walls

import openpyxl
from wall import Wall
import matplotlib.pyplot as plt
import numpy as np
from transmitter import Transmitter


def create_wall():
    # Give the location of the file to open the workbook 
    # path ="piece_rectangulaire.xlsx"
    path = "1mur.xlsx"

    # workbook object is created and activeted
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # number of rows of the excel sheet
    row_count = sheet_obj.max_row
    # Creation of the list of walls
    wall_list = []
    # filling the list of walls
    for value in sheet_obj.iter_rows(min_row=2, max_row=row_count, min_col=1, max_col=6, values_only=True):
        wall_list.append(Wall(value[4], [(value[0], value[1]), (value[2], value[3])], value[5]))
    # print(len(wall_list))
    # for i in range(0,len(wall_list)):
    # print(wall_list[i].point_list)

    return wall_list


def plot_wall(wall_list):
    # plots the walls
    for i in range(0, len(wall_list)):
        j = 0
        #
        while j < len(wall_list[i].point_list) - 1:
            #print(len(wall_list[i].point_list))
            # print(wall_list[i].point_list[j][0], wall_list[i].point_list[j][1])
            #print([wall_list[i].point_list[j][0], wall_list[i].point_list[j + 1][0]],
                  #[wall_list[i].point_list[j][1], wall_list[i].point_list[j + 1][1]])
            plt.plot([wall_list[i].point_list[j][0], wall_list[i].point_list[j + 1][0]],
                     [wall_list[i].point_list[j][1], wall_list[i].point_list[j + 1][1]], c='black')
            j = j + 2


"""def plot_wall(wall_list):
    #plots the walls
    for i in range(0, len(wall_list)):
        for j in range(0,len(wall_list[i].point_list)-1) :
            print(len(wall_list[i].point_list))
            #print(wall_list[i].point_list[j][0], wall_list[i].point_list[j][1])
            print([wall_list[i].point_list[j][0], wall_list[i].point_list[j+1][0]],
                     [wall_list[i].point_list[j][1], wall_list[i].point_list[j+1][1]])
            plt.plot([wall_list[i].point_list[j][0], wall_list[i].point_list[j+1][0]],
                     [wall_list[i].point_list[j][1], wall_list[i].point_list[j+1][1]], c='black')"""


def plott(point_list):
    """ draws lines from two points """
    x = []
    y = []
    for i in point_list:
        x.append(i[0])
        y.append(i[1])
    plt.plot(x, y)
    return 0


def power_graphical_display(string, wall_list, transmitter_list):
    # fig, ax = plt.subplots(constrained_layout=True)
    fig = plt.figure(figsize=(6, 3))
    # displays the graphic of the local power or the mean power depending on the string
    x, y, temp = np.loadtxt(string).T
    plt.scatter(x=x, y=y, c=temp, s=10, cmap=plt.cm.rainbow)
    plt.colorbar()
    # plt.gca().invert_xaxis()
    plt.gca().invert_yaxis()  # todo
    plot_wall(wall_list)
    for transmitter in transmitter_list:
        plt.scatter(transmitter.position[0], transmitter.position[1], s=30, c="black")
    plt.show()
    return 0

def ray_graphical_display(receiver, transmitter, wall_list,ray_list):
        """displays the calculted rays """
        # plt.axis([-2, 14, -2, 14])  # Todo modification des coordonnée
        # plt.gca().invert_yaxis()
        # plot_wall(self.wall_list)  # TODO
        # print("len wall list : ", len(self.wall_list))  # todo
        for ray in ray_list:
            plott(ray.point_list)
        plot_wall(wall_list)
        plt.scatter(receiver.position[0], receiver.position[1], s=40, c="red")
        plt.scatter(transmitter.position[0], transmitter.position[1], s=40, c="blue")
        plt.show()
        return 0
