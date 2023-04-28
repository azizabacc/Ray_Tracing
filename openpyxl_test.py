# Python program to read the map (excel) and generate the list of walls

import openpyxl
from Wall import Wall
from Transmitter import Transmitter

# Give the location of the file to open the workbook 
path = "C:\\Users\\aziza\\Desktop\\Télécom\\PROJET\\map.xlsx" 
# workbook object is created and activeted
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
#number of rows of the excel sheet
row_count = sheet_obj.max_row
#Creation of the list of walls
wall_list=[]
#filling the list of walls
for value in sheet_obj.iter_rows(min_row=2,max_row=row_count,min_col=1,max_col=6,values_only=True):
    wall_list.append(Wall(value[4],[(value[0],value[1]),(value[2],value[3])],value[5]))
print(len(wall_list))
for i in range(0,len(wall_list)):
    print(wall_list[i].material)


     
